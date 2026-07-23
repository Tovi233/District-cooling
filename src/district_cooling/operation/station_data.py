"""Readers for measured cold-station operation workbooks."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


DATA_SHEET_PREFIXES = ("7.",)


@dataclass(frozen=True)
class StationColumn:
    """Metadata for one measured device parameter."""

    column: str
    device: str
    parameter: str
    device_type: str


def get_device_type(device: str) -> str:
    """Return the equipment class from a station device code."""
    if device.startswith("PUMP_CW_DUAL"):
        return "PUMP_CW_DUAL"
    if device.startswith("PUMP_CHW"):
        return "PUMP_CHW"
    if device.startswith("PUMP_CW"):
        return "PUMP_CW"
    if device.startswith("PUMP_GLY"):
        return "PUMP_GLY"
    if device.startswith("PUMP_PHE"):
        return "PUMP_PHE"
    if device.startswith("SYS_TOTAL"):
        return "SYS_TOTAL"
    return device.split("_", 1)[0]


def load_station_workbook(path: str | Path) -> tuple[pd.DataFrame, list[StationColumn]]:
    """Load all data sheets from the final Xiaomeisha station workbook.

    The workbook format is:
    - row 1: device code
    - row 2: measured parameter name
    - column A: timestamp
    - row 3 onward: measured values
    """
    workbook_path = Path(path)
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    frames: list[pd.DataFrame] = []
    column_meta: dict[str, StationColumn] = {}

    for sheet in workbook.worksheets:
        if not sheet.title.startswith(DATA_SHEET_PREFIXES):
            continue

        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            continue

        device_row = list(rows[0])
        parameter_row = list(rows[1])
        headers = ["collect_time_iso"]

        for device, parameter in zip(device_row[1:], parameter_row[1:]):
            if device is None or parameter is None:
                continue
            device_code = str(device)
            parameter_name = str(parameter)
            column = f"{device_code}__{parameter_name}"
            headers.append(column)
            column_meta[column] = StationColumn(
                column=column,
                device=device_code,
                parameter=parameter_name,
                device_type=get_device_type(device_code),
            )

        frame = pd.DataFrame(rows[2:], columns=headers)
        frame.insert(0, "source_sheet", sheet.title)
        frames.append(frame)

    if not frames:
        raise ValueError(f"No station data sheets were found in {workbook_path}")

    data = pd.concat(frames, ignore_index=True)
    data["collect_time_iso"] = pd.to_datetime(data["collect_time_iso"], errors="coerce")
    data = data.dropna(subset=["collect_time_iso"]).sort_values("collect_time_iso").reset_index(drop=True)

    for column in data.columns:
        if column in {"source_sheet", "collect_time_iso"}:
            continue
        data[column] = pd.to_numeric(data[column], errors="coerce")

    return data, list(column_meta.values())
