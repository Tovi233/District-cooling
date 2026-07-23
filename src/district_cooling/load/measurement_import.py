"""Import measured building data for RC-model calibration."""

from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


MEASUREMENT_COLUMNS = [
    "time",
    "elapsed_h",
    "chiller_01_power_kw",
    "chiller_02_power_kw",
    "chiller_total_power_kw",
    "water_flow_m3_h",
    "supply_water_temperature_c",
    "return_water_temperature_c",
    "cooling_load_kw",
    "outdoor_air_temperature_c",
    "outdoor_relative_humidity_percent",
    "indoor_average_temperature_c",
    "indoor_average_relative_humidity_percent",
    "indoor_temperature_f4_c",
    "indoor_relative_humidity_f4_percent",
    "indoor_temperature_f6_c",
    "indoor_relative_humidity_f6_percent",
    "indoor_temperature_f11_c",
    "indoor_relative_humidity_f11_percent",
]

REQUIRED_MEASUREMENT_COLUMNS = {
    "time",
    "chiller_total_power_kw",
    "water_flow_m3_h",
    "supply_water_temperature_c",
    "return_water_temperature_c",
    "cooling_load_kw",
    "outdoor_air_temperature_c",
    "indoor_average_temperature_c",
}

SOURCE_HEADER_TO_COLUMN = {
    "时间": "time",
    "#01冷机电功率(kW)": "chiller_01_power_kw",
    "#02冷机电功率(kW)": "chiller_02_power_kw",
    "冷机总功率(kW)": "chiller_total_power_kw",
    "水流量(m3/h)": "water_flow_m3_h",
    "供水温度(℃)": "supply_water_temperature_c",
    "回水温度(℃)": "return_water_temperature_c",
    "冷负荷(kW)": "cooling_load_kw",
    "室外温度(℃)": "outdoor_air_temperature_c",
    "室外相对湿度(%)": "outdoor_relative_humidity_percent",
    "室内平均温度(℃)": "indoor_average_temperature_c",
    "室内平均相对湿度(%)": "indoor_average_relative_humidity_percent",
    "室内温度F4(℃)": "indoor_temperature_f4_c",
    "室内相对湿度F4(%)": "indoor_relative_humidity_f4_percent",
    "室内温度F6(℃)": "indoor_temperature_f6_c",
    "室内相对湿度F6(%)": "indoor_relative_humidity_f6_percent",
    "室内温度F11(℃)": "indoor_temperature_f11_c",
    "室内相对湿度F11(%)": "indoor_relative_humidity_f11_percent",
}

WZS2_POSITION_TO_COLUMN = {
    0: "time",
    1: "chiller_total_power_kw",
    2: "water_flow_m3_h",
    3: "supply_water_temperature_c",
    4: "return_water_temperature_c",
    5: "cooling_load_kw",
    6: "outdoor_air_temperature_c",
    7: "outdoor_relative_humidity_percent",
    8: "indoor_average_temperature_c",
    9: "indoor_average_relative_humidity_percent",
}


@dataclass(frozen=True)
class MeasurementImportSummary:
    """Summary of one imported measurement dataset."""

    output_csv_path: Path
    row_count: int
    start_time: datetime
    end_time: datetime
    time_step_s: float


def _normalize_time(value: Any) -> datetime:
    """Return an Excel time cell as a datetime."""
    if isinstance(value, datetime):
        return value
    if isinstance(value, str):
        return datetime.fromisoformat(value.strip())
    raise ValueError(f"unsupported time value: {value!r}")


def _normalize_number(value: Any) -> float | None:
    """Return a numeric cell as float, preserving blanks."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    return float(text)


def _build_header_index(headers: list[Any]) -> dict[str, int]:
    """Map normalized output column names to source column indexes."""
    index: dict[str, int] = {}
    for position, header in enumerate(headers):
        if header in SOURCE_HEADER_TO_COLUMN:
            index[SOURCE_HEADER_TO_COLUMN[header]] = position
    if REQUIRED_MEASUREMENT_COLUMNS - index.keys() and len(headers) >= 10:
        for position, column in WZS2_POSITION_TO_COLUMN.items():
            index.setdefault(column, position)
    missing = [
        column
        for column in REQUIRED_MEASUREMENT_COLUMNS
        if column not in index
    ]
    if missing:
        raise ValueError(f"missing source columns: {missing}")
    return index


def import_wzs_measurements(
    source_excel_path: str | Path,
    output_csv_path: str | Path,
    sheet_name: str | None = None,
) -> MeasurementImportSummary:
    """Import WZS Excel measurements into a normalized CSV file."""
    source_path = Path(source_excel_path)
    output_path = Path(output_csv_path)
    workbook = load_workbook(source_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        rows = worksheet.iter_rows(values_only=True)
        headers = list(next(rows))
        header_index = _build_header_index(headers)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        row_count = 0
        start_time: datetime | None = None
        end_time: datetime | None = None
        previous_time: datetime | None = None
        time_step_s: float | None = None

        with output_path.open("w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=MEASUREMENT_COLUMNS)
            writer.writeheader()
            for source_row in rows:
                if source_row[header_index["time"]] is None:
                    continue
                current_time = _normalize_time(source_row[header_index["time"]])
                if start_time is None:
                    start_time = current_time
                if previous_time is not None:
                    current_step = (current_time - previous_time).total_seconds()
                    if time_step_s is None:
                        time_step_s = current_step
                previous_time = current_time
                end_time = current_time

                output_row: dict[str, Any] = {
                    "time": current_time.isoformat(sep=" "),
                    "elapsed_h": (current_time - start_time).total_seconds() / 3600.0,
                }
                for column in MEASUREMENT_COLUMNS:
                    if column in ("time", "elapsed_h"):
                        continue
                    output_row[column] = (
                        _normalize_number(source_row[header_index[column]])
                        if column in header_index
                        else None
                    )
                writer.writerow(output_row)
                row_count += 1
    finally:
        workbook.close()

    if start_time is None or end_time is None:
        raise ValueError("no measurement rows were imported")

    return MeasurementImportSummary(
        output_csv_path=output_path,
        row_count=row_count,
        start_time=start_time,
        end_time=end_time,
        time_step_s=time_step_s or 0.0,
    )
