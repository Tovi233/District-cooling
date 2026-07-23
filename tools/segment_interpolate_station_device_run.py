"""Build segmented 15-minute tables with small-gap linear interpolation."""

from __future__ import annotations

import argparse
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_time(value: str) -> datetime:
    """Parse timestamp formats seen in station device exports."""
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return datetime.fromisoformat(value)


def split_header(header: str) -> tuple[str, str]:
    """Split DEVICE__field headers for two-row Excel headers."""
    if "__" not in header:
        return header, ""
    device, field = header.split("__", 1)
    return device, field


def try_float(value: str) -> float | None:
    """Return float value when possible, otherwise None."""
    value = value.strip()
    if not value:
        return None
    try:
        return float(value)
    except ValueError:
        return None


def format_number(value: float) -> str:
    """Format interpolated numbers without noisy trailing zeros."""
    if abs(value - round(value)) < 1.0e-10:
        return str(int(round(value)))
    return f"{value:.6f}".rstrip("0").rstrip(".")


def interpolate_row(
    before: dict[str, str],
    after: dict[str, str],
    timestamp: datetime,
    ratio: float,
    headers: list[str],
) -> dict[str, str]:
    """Linearly interpolate one inserted row between two measured rows."""
    row = {"collect_time_iso": timestamp.strftime("%Y-%m-%d %H:%M:%S")}
    for header in headers:
        if header == "collect_time_iso":
            continue
        before_value = try_float(before.get(header, ""))
        after_value = try_float(after.get(header, ""))
        if before_value is None or after_value is None:
            row[header] = ""
        else:
            row[header] = format_number(before_value + (after_value - before_value) * ratio)
    return row


def read_rows(path: Path, start_csv_line: int) -> tuple[list[str], list[dict[str, str]]]:
    """Read rows from a CSV line number, where line 1 is the header."""
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows = list(reader)

    start_data_index = max(start_csv_line - 2, 0)
    selected = rows[start_data_index:]
    for row in selected:
        timestamp = parse_time(row["collect_time_iso"])
        row["collect_time_iso"] = timestamp.strftime("%Y-%m-%d %H:%M:%S")
    return headers, selected


def build_segments(
    rows: list[dict[str, str]],
    headers: list[str],
    large_gap_missing_slots: set[int],
) -> tuple[list[list[dict[str, str]]], list[dict[str, object]]]:
    """Fill small gaps and split at selected large gaps."""
    segments: list[list[dict[str, str]]] = [[]]
    gap_records: list[dict[str, object]] = []

    for index, current in enumerate(rows):
        if not segments[-1] or segments[-1][-1] is not current:
            segments[-1].append(current)
        if index == len(rows) - 1:
            break

        following = rows[index + 1]
        current_time = parse_time(current["collect_time_iso"])
        following_time = parse_time(following["collect_time_iso"])
        gap_minutes = int((following_time - current_time).total_seconds() / 60.0)
        if gap_minutes <= 15:
            continue
        if gap_minutes % 15 != 0:
            gap_records.append(
                {
                    "from": current["collect_time_iso"],
                    "to": following["collect_time_iso"],
                    "gap_minutes": gap_minutes,
                    "action": "not_aligned_not_filled",
                }
            )
            segments.append([])
            segments[-1].append(following)
            continue

        missing_slots = gap_minutes // 15 - 1
        if missing_slots in large_gap_missing_slots:
            gap_records.append(
                {
                    "from": current["collect_time_iso"],
                    "to": following["collect_time_iso"],
                    "gap_minutes": gap_minutes,
                    "missing_15min_slots": missing_slots,
                    "action": "split_sheet_not_interpolated",
                }
            )
            segments.append([])
            segments[-1].append(following)
            continue

        for slot_index in range(1, missing_slots + 1):
            inserted_time = current_time + timedelta(minutes=15 * slot_index)
            ratio = slot_index / (missing_slots + 1)
            segments[-1].append(
                interpolate_row(current, following, inserted_time, ratio, headers)
            )
        gap_records.append(
            {
                "from": current["collect_time_iso"],
                "to": following["collect_time_iso"],
                "gap_minutes": gap_minutes,
                "missing_15min_slots": missing_slots,
                "action": "linear_interpolated",
            }
        )

    return [segment for segment in segments if segment], gap_records


def write_segment_csv(path: Path, headers: list[str], rows: list[dict[str, str]]) -> None:
    """Write one segment CSV."""
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def add_segment_sheet(
    workbook: Workbook,
    title: str,
    headers: list[str],
    rows: list[dict[str, str]],
) -> None:
    """Add one segment to an Excel workbook."""
    worksheet = workbook.create_sheet(title=title)
    device_header = ["collect_time_iso"]
    field_header = [""]
    for header in headers[1:]:
        device, field = split_header(header)
        device_header.append(device)
        field_header.append(field)
    worksheet.append(device_header)
    worksheet.append(field_header)
    for row in rows:
        worksheet.append([row.get(header, "") for header in headers])

    for header_row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    worksheet.freeze_panes = "B3"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 22
    for column_index in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 13


def process(
    input_csv: Path,
    output_dir: Path,
    start_csv_line: int,
    large_gap_missing_slots: set[int],
) -> dict[str, object]:
    """Create segmented interpolated outputs."""
    headers, rows = read_rows(input_csv, start_csv_line)
    segments, gap_records = build_segments(rows, headers, large_gap_missing_slots)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    csv_outputs: list[str] = []
    for index, segment in enumerate(segments, start=1):
        sheet_name = f"sheet{index}"
        add_segment_sheet(workbook, sheet_name, headers, segment)
        csv_path = output_dir / f"station_device_run_{sheet_name}_interpolated.csv"
        write_segment_csv(csv_path, headers, segment)
        csv_outputs.append(str(csv_path))

    xlsx_path = output_dir / "station_device_run_segmented_interpolated_from_row33.xlsx"
    workbook.save(xlsx_path)

    summary = {
        "input_csv": str(input_csv),
        "output_xlsx": str(xlsx_path),
        "output_csv_files": csv_outputs,
        "start_csv_line": start_csv_line,
        "segment_count": len(segments),
        "segments": [
            {
                "sheet": f"sheet{index}",
                "row_count": len(segment),
                "start": segment[0]["collect_time_iso"],
                "end": segment[-1]["collect_time_iso"],
            }
            for index, segment in enumerate(segments, start=1)
        ],
        "gap_records": gap_records,
    }
    summary_path = output_dir / "station_device_run_segmented_interpolated_summary.json"
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    summary["summary_json"] = str(summary_path)
    return summary


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=Path(
            "data/processed/station_device_run_wide/"
            "station_device_run_all_devices.csv"
        ),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/processed/station_device_run_wide/segmented_interpolated"),
    )
    parser.add_argument("--start-csv-line", type=int, default=33)
    parser.add_argument(
        "--large-gap-missing-slots",
        type=int,
        nargs="*",
        default=[65, 7],
    )
    args = parser.parse_args()
    summary = process(
        input_csv=args.input_csv,
        output_dir=args.output_dir,
        start_csv_line=args.start_csv_line,
        large_gap_missing_slots=set(args.large_gap_missing_slots),
    )
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
