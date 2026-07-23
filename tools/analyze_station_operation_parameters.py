"""Analyze station operation control ranges and typical-day parameters."""

from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd


DATA_SHEET_PREFIXES = ("7.",)
FIELD_CN = {
    "status": "运行状态",
    "power_kw": "功率",
    "freq_hz": "频率",
    "chw_in_temp": "冷冻水/载冷剂进口温度",
    "chw_out_temp": "冷冻水/载冷剂出口温度",
    "cw_in_temp": "冷却水进口温度",
    "cw_out_temp": "冷却水出口温度",
    "setpoint_temp": "出水温度设定值",
    "inventory_rt": "蓄冰量",
    "flow_m3h": "系统流量",
    "cooling_load": "系统冷量",
}
DEVICE_CN = {
    "CH": "冷水机组",
    "CT": "冷却塔",
    "HE": "换热器",
    "ICE": "蓄冰装置",
    "PUMP_CHW": "冷冻水泵",
    "PUMP_CW": "冷却水泵",
    "PUMP_CW_DUAL": "双工况冷却水泵",
    "PUMP_GLY": "乙二醇泵",
    "PUMP_PHE": "板换侧水泵",
    "SYS_TOTAL": "系统总表",
}


@dataclass(frozen=True)
class ColumnInfo:
    column: str
    device: str
    field: str
    device_type: str


def device_type(device: str) -> str:
    """Return HVAC device class from device code."""
    if device.startswith("PUMP_CW_DUAL"):
        return "PUMP_CW_DUAL"
    if device.startswith("PUMP_CHW"):
        return "PUMP_CHW"
    if device.startswith("PUMP_CW"):
        return "PUMP_CW"
    if device.startswith("PUMP_GLY"):
        return "PUMP_GLY"
    if device.startswith("PUMP_PHE"):
        return "PUMP_PHE"
    if device.startswith("SYS_TOTAL"):
        return "SYS_TOTAL"
    return device.split("_", 1)[0]


def load_workbook_data(path: Path) -> tuple[pd.DataFrame, list[ColumnInfo]]:
    """Load the three data sheets into one DataFrame."""
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    frames: list[pd.DataFrame] = []
    column_infos: dict[str, ColumnInfo] = {}
    for sheet in workbook.worksheets:
        if not sheet.title.startswith(DATA_SHEET_PREFIXES):
            continue
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) < 3:
            continue
        device_row = list(rows[0])
        field_row = list(rows[1])
        headers = ["collect_time_iso"]
        for device, field in zip(device_row[1:], field_row[1:]):
            column = f"{device}__{field}"
            headers.append(column)
            column_infos[column] = ColumnInfo(
                column=column,
                device=str(device),
                field=str(field),
                device_type=device_type(str(device)),
            )
        frame = pd.DataFrame(rows[2:], columns=headers)
        frame.insert(0, "source_sheet", sheet.title)
        frames.append(frame)
    if not frames:
        raise ValueError(f"no data sheets found in {path}")

    data = pd.concat(frames, ignore_index=True)
    data["collect_time_iso"] = pd.to_datetime(data["collect_time_iso"])
    for column in data.columns:
        if column in {"source_sheet", "collect_time_iso"}:
            continue
        data[column] = pd.to_numeric(data[column], errors="coerce")

    # Frequency -59 is a device data flag, not a physical operating frequency.
    freq_columns = [column for column, info in column_infos.items() if info.field == "freq_hz"]
    for column in freq_columns:
        data.loc[data[column] < 0, column] = pd.NA
    return data, list(column_infos.values())


def parameter_ranges(data: pd.DataFrame, infos: list[ColumnInfo]) -> pd.DataFrame:
    """Return range statistics for every device-parameter column."""
    records: list[dict[str, Any]] = []
    for info in infos:
        series = pd.to_numeric(data[info.column], errors="coerce")
        valid = series.dropna()
        records.append(
            {
                "设备类型": info.device_type,
                "暖通中文名称": DEVICE_CN.get(info.device_type, info.device_type),
                "设备编号": info.device,
                "参数": info.field,
                "参数中文": FIELD_CN.get(info.field, info.field),
                "有效点数": int(valid.count()),
                "缺失点数": int(series.isna().sum()),
                "最小值": valid.min() if not valid.empty else pd.NA,
                "最大值": valid.max() if not valid.empty else pd.NA,
                "平均值": valid.mean() if not valid.empty else pd.NA,
                "中位数": valid.median() if not valid.empty else pd.NA,
                "P05": valid.quantile(0.05) if not valid.empty else pd.NA,
                "P95": valid.quantile(0.95) if not valid.empty else pd.NA,
            }
        )
    return pd.DataFrame(records)


def device_type_ranges(ranges: pd.DataFrame) -> pd.DataFrame:
    """Aggregate ranges by device type and parameter."""
    grouped = (
        ranges.groupby(["设备类型", "暖通中文名称", "参数", "参数中文"], dropna=False)
        .agg(
            设备数量=("设备编号", "nunique"),
            测点数=("设备编号", "count"),
            有效点数=("有效点数", "sum"),
            缺失点数=("缺失点数", "sum"),
            最小值=("最小值", "min"),
            最大值=("最大值", "max"),
            平均值=("平均值", "mean"),
            P05=("P05", "min"),
            P95=("P95", "max"),
        )
        .reset_index()
    )
    return grouped


def add_operation_features(data: pd.DataFrame) -> pd.DataFrame:
    """Add station-level features for typical-day analysis."""
    out = data[["source_sheet", "collect_time_iso"]].copy()
    out["date"] = out["collect_time_iso"].dt.date.astype(str)
    out["hour"] = out["collect_time_iso"].dt.hour + out["collect_time_iso"].dt.minute / 60.0
    sys_cols = {
        "总功率_kW": "SYS_TOTAL__power_kw",
        "总流量_m3h": "SYS_TOTAL__flow_m3h",
        "总冷量": "SYS_TOTAL__cooling_load",
        "蓄冰量": "ICE_01__inventory_rt",
    }
    for target, source in sys_cols.items():
        out[target] = pd.to_numeric(data[source], errors="coerce") if source in data else pd.NA
    out["单位冷量功率"] = out["总功率_kW"] / out["总冷量"].replace(0, pd.NA)
    out["蓄冰量变化_每15min"] = out["蓄冰量"].diff()

    ch_status_cols = [column for column in data.columns if column.startswith("CH_") and column.endswith("__status")]
    ct_power_cols = [column for column in data.columns if column.startswith("CT_") and column.endswith("__power_kw")]
    pump_power_cols = [column for column in data.columns if column.startswith("PUMP_") and column.endswith("__power_kw")]
    out["冷机运行台数"] = data[ch_status_cols].fillna(0).gt(0.5).sum(axis=1)
    out["冷却塔运行台数"] = data[ct_power_cols].fillna(0).gt(1.0).sum(axis=1)
    out["水泵运行台数"] = data[pump_power_cols].fillna(0).gt(1.0).sum(axis=1)
    return out


def daily_summary(features: pd.DataFrame) -> pd.DataFrame:
    """Return daily operation parameter summaries."""
    grouped = (
        features.groupby("date", dropna=False)
        .agg(
            时间点数=("collect_time_iso", "count"),
            起始时间=("collect_time_iso", "min"),
            结束时间=("collect_time_iso", "max"),
            平均冷量=("总冷量", "mean"),
            最大冷量=("总冷量", "max"),
            最小冷量=("总冷量", "min"),
            平均总功率_kW=("总功率_kW", "mean"),
            最大总功率_kW=("总功率_kW", "max"),
            平均流量_m3h=("总流量_m3h", "mean"),
            最大流量_m3h=("总流量_m3h", "max"),
            平均单位冷量功率=("单位冷量功率", "mean"),
            平均冷机运行台数=("冷机运行台数", "mean"),
            最大冷机运行台数=("冷机运行台数", "max"),
            平均冷却塔运行台数=("冷却塔运行台数", "mean"),
            最大冷却塔运行台数=("冷却塔运行台数", "max"),
            平均水泵运行台数=("水泵运行台数", "mean"),
            最大水泵运行台数=("水泵运行台数", "max"),
            蓄冰量起始=("蓄冰量", "first"),
            蓄冰量结束=("蓄冰量", "last"),
            蓄冰量最小=("蓄冰量", "min"),
            蓄冰量最大=("蓄冰量", "max"),
        )
        .reset_index()
    )
    grouped["蓄冰量净变化"] = grouped["蓄冰量结束"] - grouped["蓄冰量起始"]
    grouped["完整度_按96点"] = grouped["时间点数"] / 96.0
    return grouped


def typical_day(daily: pd.DataFrame) -> pd.DataFrame:
    """Select the most representative available day by average cooling load."""
    complete = daily[daily["完整度_按96点"] >= 0.8].copy()
    if complete.empty:
        complete = daily.copy()
    median_load = complete["平均冷量"].median()
    complete["与中位日冷量差"] = (complete["平均冷量"] - median_load).abs()
    selected = complete.sort_values(["与中位日冷量差", "时间点数"], ascending=[True, False]).head(1)
    return selected


def write_report(
    output_path: Path,
    ranges: pd.DataFrame,
    type_ranges: pd.DataFrame,
    daily: pd.DataFrame,
    selected_typical_day: pd.DataFrame,
    features: pd.DataFrame,
) -> None:
    """Write a concise Markdown analysis report."""
    total_rows = len(features)
    start = features["collect_time_iso"].min()
    end = features["collect_time_iso"].max()
    typical_date = selected_typical_day.iloc[0]["date"]
    lines = [
        "# 小梅沙冷站运行调控参数范围与典型日参数",
        "",
        "## 数据概况",
        "",
        f"- 数据时间范围: {start} 至 {end}",
        f"- 总时间点数: {total_rows}",
        f"- 设备参数测点数: {len(ranges)}",
        f"- 可用自然日数: {daily['date'].nunique()}",
        f"- 典型日建议: {typical_date}",
        "",
        "## 系统级调控参数范围",
        "",
    ]
    system_rows = type_ranges[type_ranges["设备类型"] == "SYS_TOTAL"]
    for _, row in system_rows.iterrows():
        lines.append(
            f"- {row['参数中文']}({row['参数']}): "
            f"{row['最小值']:.3f} ~ {row['最大值']:.3f}, "
            f"均值 {row['平均值']:.3f}"
        )
    lines.extend(["", "## 主要设备调控范围", ""])
    for _, row in type_ranges.iterrows():
        if row["设备类型"] == "SYS_TOTAL":
            continue
        if row["参数"] not in {"power_kw", "freq_hz", "setpoint_temp", "status", "inventory_rt"}:
            continue
        lines.append(
            f"- {row['暖通中文名称']} {row['参数中文']}({row['参数']}): "
            f"{row['最小值']:.3f} ~ {row['最大值']:.3f}, "
            f"测点 {int(row['测点数'])} 个"
        )
    lines.extend(["", "## 典型日运行参数", ""])
    selected = selected_typical_day.iloc[0]
    for key in [
        "时间点数",
        "平均冷量",
        "最大冷量",
        "平均总功率_kW",
        "最大总功率_kW",
        "平均流量_m3h",
        "最大流量_m3h",
        "平均冷机运行台数",
        "最大冷机运行台数",
        "平均冷却塔运行台数",
        "最大冷却塔运行台数",
        "平均水泵运行台数",
        "最大水泵运行台数",
        "蓄冰量起始",
        "蓄冰量结束",
        "蓄冰量净变化",
    ]:
        value = selected[key]
        lines.append(f"- {key}: {value}")
    output_path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def analyze(source: Path, output_dir: Path) -> dict[str, str]:
    """Run analysis and write output files."""
    output_dir.mkdir(parents=True, exist_ok=True)
    data, infos = load_workbook_data(source)
    ranges = parameter_ranges(data, infos)
    type_ranges = device_type_ranges(ranges)
    features = add_operation_features(data)
    daily = daily_summary(features)
    selected_typical_day = typical_day(daily)
    typical_date = selected_typical_day.iloc[0]["date"]
    typical_profile = features[features["date"] == typical_date].copy()

    paths = {
        "参数范围明细": output_dir / "station_parameter_ranges.csv",
        "设备类型调控范围": output_dir / "station_device_type_control_ranges.csv",
        "日运行参数汇总": output_dir / "station_daily_operation_summary.csv",
        "典型日运行参数": output_dir / "station_typical_day_profile.csv",
        "特征时间序列": output_dir / "station_operation_features_timeseries.csv",
        "Markdown汇报": output_dir / "station_operation_parameter_report.md",
    }
    ranges.to_csv(paths["参数范围明细"], index=False, encoding="utf-8-sig")
    type_ranges.to_csv(paths["设备类型调控范围"], index=False, encoding="utf-8-sig")
    daily.to_csv(paths["日运行参数汇总"], index=False, encoding="utf-8-sig")
    typical_profile.to_csv(paths["典型日运行参数"], index=False, encoding="utf-8-sig")
    features.to_csv(paths["特征时间序列"], index=False, encoding="utf-8-sig")
    write_report(
        paths["Markdown汇报"],
        ranges=ranges,
        type_ranges=type_ranges,
        daily=daily,
        selected_typical_day=selected_typical_day,
        features=features,
    )
    return {name: str(path) for name, path in paths.items()}


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--source",
        type=Path,
        default=Path("data/小梅沙.xlsx"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/processed/xiaomeisha_operation_analysis"),
    )
    args = parser.parse_args()
    outputs = analyze(args.source, args.output_dir)
    for name, path in outputs.items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()
