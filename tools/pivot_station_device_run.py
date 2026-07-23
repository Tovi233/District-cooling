"""Create time-by-device wide tables from normalized station device run data."""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


DEFAULT_VALUE_COLUMNS = [
    "status",
    "power_kw",
    "freq_hz",
    "chw_in_temp",
    "chw_out_temp",
    "cw_in_temp",
    "cw_out_temp",
    "setpoint_temp",
    "valve_open",
    "inventory_rt",
    "flow_m3h",
    "cooling_load",
]


def read_rows(path: Path) -> list[dict[str, str]]:
    """Read normalized long-form rows."""
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        return list(csv.DictReader(file))


def pivot_rows(
    rows: list[dict[str, str]],
    value_column: str,
) -> tuple[list[str], list[str], list[dict[str, str]]]:
    """Return timestamp rows with device-id columns for one value field."""
    timestamps = sorted({row["collect_time_iso"] for row in rows if row["collect_time_iso"]})
    devices = sorted({row["device_id"] for row in rows if row["device_id"]})
    values: dict[tuple[str, str], str] = {}
    duplicate_counts: dict[tuple[str, str], int] = {}

    for row in rows:
        timestamp = row["collect_time_iso"]
        device = row["device_id"]
        if not timestamp or not device:
            continue
        key = (timestamp, device)
        value = row.get(value_column, "")
        if key in values and value:
            duplicate_counts[key] = duplicate_counts.get(key, 1) + 1
        if value or key not in values:
            values[key] = value

    table: list[dict[str, str]] = []
    for timestamp in timestamps:
        record = {"collect_time_iso": timestamp}
        for device in devices:
            record[device] = values.get((timestamp, device), "")
        table.append(record)
    return timestamps, devices, table


def write_csv(path: Path, devices: list[str], table: list[dict[str, str]]) -> None:
    """Write one wide CSV."""
    fieldnames = ["collect_time_iso", *devices]
    with path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(table)


def write_excel_sheet(
    workbook: Workbook,
    sheet_name: str,
    devices: list[str],
    table: list[dict[str, str]],
) -> None:
    """Add one pivot table sheet."""
    worksheet = workbook.create_sheet(title=sheet_name[:31])
    headers = ["collect_time_iso", *devices]
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")

    for record in table:
        worksheet.append([record.get(header, "") for header in headers])

    worksheet.freeze_panes = "B2"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 22
    for index in range(2, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 14


def build_pivots(
    normalized_csv: Path,
    output_dir: Path,
    value_columns: list[str],
) -> dict[str, str]:
    """Build CSV and XLSX wide tables."""
    rows = read_rows(normalized_csv)
    output_dir.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    workbook.remove(workbook.active)
    output_paths: dict[str, str] = {}

    for value_column in value_columns:
        _, devices, table = pivot_rows(rows, value_column)
        csv_path = output_dir / f"station_device_run_wide_{value_column}.csv"
        write_csv(csv_path, devices, table)
        write_excel_sheet(workbook, value_column, devices, table)
        output_paths[value_column] = str(csv_path)

    xlsx_path = output_dir / "station_device_run_wide_tables.xlsx"
    workbook.save(xlsx_path)
    output_paths["excel_workbook"] = str(xlsx_path)
    return output_paths


def build_all_devices_table(
    normalized_csv: Path,
    output_dir: Path,
    value_columns: list[str],
    drop_timestamps: set[str] | None = None,
) -> dict[str, str]:
    """Build one table containing all devices and all measured value columns."""
    rows = read_rows(normalized_csv)
    drop_timestamps = drop_timestamps or set()
    rows = [
        row
        for row in rows
        if row.get("collect_time_iso", "") not in drop_timestamps
    ]
    output_dir.mkdir(parents=True, exist_ok=True)
    timestamps = sorted({row["collect_time_iso"] for row in rows if row["collect_time_iso"]})
    devices = sorted({row["device_id"] for row in rows if row["device_id"]})

    values: dict[tuple[str, str, str], str] = {}
    non_empty_device_columns: set[tuple[str, str]] = set()
    for row in rows:
        timestamp = row["collect_time_iso"]
        device = row["device_id"]
        if not timestamp or not device:
            continue
        for value_column in value_columns:
            value = row.get(value_column, "")
            if value:
                values[(timestamp, device, value_column)] = value
                non_empty_device_columns.add((device, value_column))

    device_value_columns = [
        (device, value_column)
        for device in devices
        for value_column in value_columns
        if (device, value_column) in non_empty_device_columns
    ]

    csv_headers = [
        "collect_time_iso",
        *[
            f"{device}__{value_column}"
            for device, value_column in device_value_columns
        ],
    ]
    csv_path = output_dir / "station_device_run_all_devices.csv"
    with csv_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.writer(file)
        writer.writerow(csv_headers)
        for timestamp in timestamps:
            writer.writerow(
                [
                    timestamp,
                    *[
                        values.get((timestamp, device, value_column), "")
                        for device, value_column in device_value_columns
                    ],
                ]
            )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "all_devices"
    worksheet.append(
        [
            "collect_time_iso",
            *[
                device
                for device, _ in device_value_columns
            ],
        ]
    )
    worksheet.append(
        [
            "",
            *[
                value_column
                for _, value_column in device_value_columns
            ],
        ]
    )
    for row_index, timestamp in enumerate(timestamps, start=3):
        worksheet.cell(row=row_index, column=1, value=timestamp)
        column_index = 2
        for device, value_column in device_value_columns:
            worksheet.cell(
                row=row_index,
                column=column_index,
                value=values.get((timestamp, device, value_column), ""),
            )
            column_index += 1

    for row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    worksheet.freeze_panes = "B3"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 22
    for index in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(index)].width = 13

    xlsx_path = output_dir / "station_device_run_all_devices.xlsx"
    workbook.save(xlsx_path)
    return {
        "csv": str(csv_path),
        "excel": str(xlsx_path),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--normalized-csv",
        type=Path,
        default=Path("data/processed/station_device_run_normalized.csv"),
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=Path("data/processed/station_device_run_wide"),
    )
    parser.add_argument(
        "--columns",
        nargs="*",
        default=DEFAULT_VALUE_COLUMNS,
    )
    parser.add_argument(
        "--single-table",
        action="store_true",
        help="write one table with all devices and all value columns",
    )
    parser.add_argument(
        "--drop-timestamp",
        action="append",
        default=[],
        help="collect_time_iso timestamp to exclude; can be used multiple times",
    )
    args = parser.parse_args()
    outputs = (
        build_all_devices_table(
            args.normalized_csv,
            args.output_dir,
            args.columns,
            set(args.drop_timestamp),
        )
        if args.single_table
        else build_pivots(args.normalized_csv, args.output_dir, args.columns)
    )
    for name, path in outputs.items():
        print(f"{name}: {path}")


if __name__ == "__main__":
    main()
