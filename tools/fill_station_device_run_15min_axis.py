"""Create a continuous 15-minute time-axis table from the all-device wide table."""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def parse_time(value: str) -> datetime:
    """Parse known timestamp formats in exported CSV files."""
    value = value.strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M", "%Y/%m/%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            pass
    return datetime.fromisoformat(value)


def floor_to_quarter_hour(value: datetime) -> datetime:
    """Floor timestamp to a clock quarter hour."""
    minute = (value.minute // 15) * 15
    return value.replace(minute=minute, second=0, microsecond=0)


def ceil_to_quarter_hour(value: datetime) -> datetime:
    """Ceil timestamp to a clock quarter hour."""
    floored = floor_to_quarter_hour(value)
    if floored == value.replace(second=0, microsecond=0):
        return floored
    return floored + timedelta(minutes=15)


def device_and_field(header: str) -> tuple[str, str]:
    """Split a CSV column name like DEVICE__field."""
    if "__" not in header:
        return header, ""
    return tuple(header.split("__", 1))  # type: ignore[return-value]


def read_table(path: Path) -> tuple[list[str], dict[datetime, dict[str, str]]]:
    """Read the current all-device wide CSV."""
    with path.open("r", encoding="utf-8-sig", newline="") as file:
        reader = csv.DictReader(file)
        headers = reader.fieldnames or []
        rows_by_time: dict[datetime, dict[str, str]] = {}
        for row in reader:
            timestamp = parse_time(row["collect_time_iso"])
            rows_by_time[timestamp.replace(second=0, microsecond=0)] = row
    return headers, rows_by_time


def continuous_times(start: datetime, end: datetime) -> list[datetime]:
    """Return a continuous 15-minute time axis."""
    times: list[datetime] = []
    current = start
    while current <= end:
        times.append(current)
        current += timedelta(minutes=15)
    return times


def non_empty_headers(
    headers: list[str],
    table_rows: list[dict[str, str]],
) -> list[str]:
    """Drop columns that are empty after the 15-minute alignment."""
    kept = ["collect_time_iso"]
    for header in headers:
        if header == "collect_time_iso":
            continue
        if any(row.get(header, "") for row in table_rows):
            kept.append(header)
    return kept


def build_continuous_table(
    input_csv: Path,
    output_csv: Path,
    output_xlsx: Path,
) -> dict[str, int | str]:
    """Build a new continuous 15-minute CSV and XLSX table."""
    headers, rows_by_time = read_table(input_csv)
    if not rows_by_time:
        raise ValueError("input table has no data rows")

    start = floor_to_quarter_hour(min(rows_by_time))
    end = ceil_to_quarter_hour(max(rows_by_time))
    timeline = continuous_times(start, end)
    raw_headers = [header for header in headers if header]
    full_rows: list[dict[str, str]] = []
    matched_count = 0
    blank_inserted_count = 0

    for timestamp in timeline:
        source = rows_by_time.get(timestamp)
        if source is None:
            blank_inserted_count += 1
            row = {"collect_time_iso": timestamp.strftime("%Y-%m-%d %H:%M:%S")}
        else:
            matched_count += 1
            row = {
                header: source.get(header, "")
                for header in raw_headers
            }
            row["collect_time_iso"] = timestamp.strftime("%Y-%m-%d %H:%M:%S")
        full_rows.append(row)

    output_headers = non_empty_headers(raw_headers, full_rows)
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with output_csv.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=output_headers)
        writer.writeheader()
        for row in full_rows:
            writer.writerow({header: row.get(header, "") for header in output_headers})

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "15min_filled"
    first_header_row = ["collect_time_iso"]
    second_header_row = [""]
    for header in output_headers[1:]:
        device, field = device_and_field(header)
        first_header_row.append(device)
        second_header_row.append(field)
    worksheet.append(first_header_row)
    worksheet.append(second_header_row)
    for row in full_rows:
        worksheet.append([row.get(header, "") for header in output_headers])

    for header_row in worksheet.iter_rows(min_row=1, max_row=2):
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9EAF7")
    worksheet.freeze_panes = "B3"
    worksheet.auto_filter.ref = worksheet.dimensions
    worksheet.column_dimensions["A"].width = 22
    for column_index in range(2, worksheet.max_column + 1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = 13
    workbook.save(output_xlsx)

    return {
        "input_csv": str(input_csv),
        "output_csv": str(output_csv),
        "output_xlsx": str(output_xlsx),
        "start": start.strftime("%Y-%m-%d %H:%M:%S"),
        "end": end.strftime("%Y-%m-%d %H:%M:%S"),
        "timeline_rows": len(timeline),
        "matched_existing_rows": matched_count,
        "blank_inserted_rows": blank_inserted_count,
        "input_columns": len(raw_headers),
        "output_columns": len(output_headers),
        "dropped_empty_columns": len(raw_headers) - len(output_headers),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input-csv",
        type=Path,
        default=Path(
            "data/processed/station_device_run_wide/"
            "station_device_run_all_devices.csv"
        ),
    )
    parser.add_argument(
        "--output-csv",
        type=Path,
        default=Path(
            "data/processed/station_device_run_wide/"
            "station_device_run_all_devices_15min_filled.csv"
        ),
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=Path(
            "data/processed/station_device_run_wide/"
            "station_device_run_all_devices_15min_filled.xlsx"
        ),
    )
    args = parser.parse_args()
    summary = build_continuous_table(args.input_csv, args.output_csv, args.output_xlsx)
    for key, value in summary.items():
        print(f"{key}: {value}")


if __name__ == "__main__":
    main()
